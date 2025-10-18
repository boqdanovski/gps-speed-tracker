from http.server import BaseHTTPRequestHandler
import json
import os
from datetime import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Создаем директорию для данных
DATA_DIR = '/tmp/speed_data'
ALL_DEVICES_FILE = os.path.join(DATA_DIR, 'all_devices.txt')
os.makedirs(DATA_DIR, exist_ok=True)

# Функция для получения московского времени
def get_moscow_time():
    moscow_tz = pytz.timezone('Europe/Moscow')
    return datetime.now(moscow_tz)

def create_excel_file():
    """Создает или обновляет Excel файл с данными о скорости всех устройств"""
    try:
        excel_file = os.path.join(DATA_DIR, 'gps_speed_data.xlsx')
        
        # Проверяем, существует ли файл
        if os.path.exists(excel_file):
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            # Создаем новый файл
            wb = Workbook()
            ws = wb.active
            ws.title = "GPS Speed Data"
            
            # Заголовки
            headers = ["Устройство", "Скорость (км/ч)", "Время"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Настройка ширины колонок
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
        
        # Получаем все файлы устройств
        device_files = []
        if os.path.exists(DATA_DIR):
            device_files = [f for f in os.listdir(DATA_DIR) if f.startswith('device_') and f.endswith('.txt') and not f.endswith('_log.txt')]
        
        # Очищаем старые данные (кроме заголовков)
        if ws.max_row > 1:
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)
        
        # Заполняем данными
        row = 2
        for filename in sorted(device_files):
            device_name = filename.replace('device_', '').replace('.txt', '').replace('_', ' ')
            filepath = os.path.join(DATA_DIR, filename)
            
            try:
                with open(filepath, 'r') as f:
                    content = f.read().strip()
                
                lines = content.split('\n')
                if len(lines) >= 2:
                    speed = lines[0]
                    timestamp_str = lines[1]
                    
                    # Парсим время
                    try:
                        dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                        time_only = dt.strftime('%H:%M:%S')
                    except:
                        time_only = timestamp_str
                    
                    # Записываем данные
                    ws.cell(row=row, column=1, value=device_name)
                    ws.cell(row=row, column=2, value=float(speed) if speed.replace('.', '').isdigit() else speed)
                    ws.cell(row=row, column=3, value=time_only)
                    
                    row += 1
                    
            except Exception as e:
                print(f"Ошибка обработки файла {filename}: {e}")
                continue
        
        # Сохраняем файл
        wb.save(excel_file)
        
        print(f"✅ Excel файл обновлен: {excel_file}")
        return excel_file
        
    except Exception as e:
        print(f"❌ Ошибка создания Excel файла: {e}")
        return None

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        # Получаем IP клиента
        client_ip = self.headers.get('X-Forwarded-For', 'unknown').split(',')[0].strip()
        if client_ip == 'unknown':
            client_ip = self.headers.get('X-Real-IP', 'unknown')
        
        # Получаем название устройства
        device_name = self.headers.get('X-Device-Name', client_ip)
        safe_name = device_name.replace('.', '_').replace(':', '_').replace(' ', '_')
        device_file = os.path.join(DATA_DIR, f'device_{safe_name}.txt')
        device_log_file = os.path.join(DATA_DIR, f'device_{safe_name}_log.txt')

        # Читаем данные
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        speed_data = post_data.decode('utf-8').strip()

        now = get_moscow_time()
        timestamp = now.strftime('%Y-%m-%d %H:%M:%S')

        print(f'📥 Получена скорость от {device_name} ({client_ip}): {speed_data} км/ч в {timestamp}')

        # Сохраняем данные с временной меткой
        with open(device_file, 'w') as f:
            f.write(f"{speed_data}\n{timestamp}")

        with open(device_log_file, 'a') as f:
            f.write(f'{timestamp} - {speed_data} км/ч\n')

        with open(ALL_DEVICES_FILE, 'a') as f:
            f.write(f'{timestamp} - {device_name} ({client_ip}) - {speed_data} км/ч\n')

        # Обновляем Excel файл
        try:
            create_excel_file()
        except Exception as e:
            print(f"⚠️ Ошибка обновления Excel: {e}")

        # Отправляем ответ
        self.send_response(200)
        self.send_header('Content-type', 'text/plain')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        self.end_headers()
        self.wfile.write(f'Speed updated for {device_name}: {speed_data} km/h'.encode())

    def handle_file_download(self):
        try:
            filename = self.path.replace('/download/', '')
            
            # Разрешенные файлы для скачивания
            allowed_files = [
                'all_devices.txt',
                'GPS-Speed-69F-v3.0-With-Remote-Restart.apk',
                'restart_signal.txt',
                'gps_speed_data.xlsx'
            ]
            
            # Проверяем, что файл разрешен для скачивания
            if not (filename.startswith('device_') and (filename.endswith('.txt') or filename.endswith('_log.txt'))) and filename not in allowed_files:
                self.send_error(404, "File not found")
                return
            
            # Определяем путь к файлу
            if filename == 'GPS-Speed-69F-v3.0-With-Remote-Restart.apk':
                # APK файл находится в корне проекта
                filepath = filename
            else:
                # Остальные файлы в DATA_DIR
                filepath = os.path.join(DATA_DIR, filename)
            
            if not os.path.exists(filepath):
                self.send_error(404, "File not found")
                return
            
            # Определяем тип контента
            if filename.endswith('.apk'):
                content_type = 'application/vnd.android.package-archive'
                # Читаем APK как бинарный файл
                with open(filepath, 'rb') as f:
                    content = f.read()
            elif filename.endswith('.xlsx'):
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                # Читаем Excel как бинарный файл
                with open(filepath, 'rb') as f:
                    content = f.read()
            else:
                content_type = 'text/plain; charset=utf-8'
                # Читаем текстовые файлы
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read().encode('utf-8')
            
            self.send_response(200)
            self.send_header('Content-Type', content_type)
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
            self.end_headers()
            self.wfile.write(content)
            
        except Exception as e:
            print(f'❌ Ошибка при скачивании файла: {e}')
            self.send_error(500, "Internal server error")

    def handle_cleanup(self):
        try:
            cleaned_files = []
            if os.path.exists(DATA_DIR):
                device_files = [f for f in os.listdir(DATA_DIR) if f.startswith('device_') and f.endswith('.txt') and not f.endswith('_log.txt')]
                
                for filename in device_files:
                    filepath = os.path.join(DATA_DIR, filename)
                    try:
                        os.remove(filepath)
                        cleaned_files.append(filename)
                        
                        log_filename = filename.replace('.txt', '_log.txt')
                        log_filepath = os.path.join(DATA_DIR, log_filename)
                        if os.path.exists(log_filepath):
                            os.remove(log_filepath)
                            cleaned_files.append(log_filename)
                            
                    except Exception as e:
                        print(f'❌ Ошибка удаления файла {filename}: {e}')
            
            if os.path.exists(ALL_DEVICES_FILE):
                with open(ALL_DEVICES_FILE, 'w') as f:
                    f.write('')
                cleaned_files.append('all_devices.txt')
            
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
            self.end_headers()
            
            html = f'''<!DOCTYPE html>
<html><head><title>Очистка данных</title></head>
<body>
<h1>🧹 Очистка завершена!</h1>
<p>Удалено файлов: {len(cleaned_files)}</p>
<p><a href="/">← Вернуться к мониторингу</a></p>
</body></html>'''
            
            self.wfile.write(html.encode('utf-8'))
            print(f'🧹 Очистка завершена. Удалено файлов: {len(cleaned_files)}')
            
        except Exception as e:
            print(f'❌ Ошибка при очистке: {e}')
            self.send_error(500, "Internal server error")

    def handle_restart_tracking(self):
        """Обработка команды перезапуска tracking"""
        try:
            # Создаем специальный файл-сигнал для перезапуска
            restart_file = os.path.join(DATA_DIR, 'restart_signal.txt')
            with open(restart_file, 'w') as f:
                f.write(f"RESTART_TRACKING\n{get_moscow_time().strftime('%Y-%m-%d %H:%M:%S')}")
            
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
            self.end_headers()
            
            html_content = f'''<!DOCTYPE html>
<html>
<head>
    <title>🔄 Перезапуск Tracking - 69F СКОРОСТЬ</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body {{ 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; 
            margin: 0; 
            padding: 20px; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }}
        .container {{
            max-width: 600px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .content {{
            padding: 30px;
        }}
        .success {{
            background: #d4edda;
            border: 1px solid #c3e6cb;
            color: #155724;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .info {{
            background: #d1ecf1;
            border: 1px solid #bee5eb;
            color: #0c5460;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .back-link {{
            display: inline-block;
            background: #007bff;
            color: white;
            padding: 10px 20px;
            text-decoration: none;
            border-radius: 5px;
            margin-top: 20px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🔄 Перезапуск Tracking</h1>
            <p>Команда отправлена устройствам</p>
        </div>
        <div class="content">
            <div class="success">
                ✅ <strong>Команда перезапуска отправлена!</strong><br>
                Время: {get_moscow_time().strftime('%Y-%m-%d %H:%M:%S')} (МСК)
            </div>
            
            <div class="info">
                <h3>📱 Что происходит:</h3>
                <ul>
                    <li>Создан файл-сигнал для перезапуска</li>
                    <li>Приложения проверят этот файл при следующей отправке</li>
                    <li>Tracking будет автоматически перезапущен</li>
                    <li>Новые данные появятся через 1-2 минуты</li>
                </ul>
            </div>
            
            <div class="info">
                <h3>🔗 Файл-сигнал:</h3>
                <code>https://gps-speed-tracker.vercel.app/download/restart_signal.txt</code>
            </div>
            
            <a href="/" class="back-link">← Вернуться к мониторингу</a>
        </div>
    </div>
</body>
</html>'''
            
            self.wfile.write(html_content.encode('utf-8'))
            print(f'🔄 Команда перезапуска отправлена в {get_moscow_time().strftime("%Y-%m-%d %H:%M:%S")}')
            
        except Exception as e:
            print(f'❌ Ошибка при отправке команды перезапуска: {e}')
            self.send_error(500, "Internal server error")

    def handle_create_excel(self):
        """Обработка создания Excel файла"""
        try:
            # Создаем Excel файл
            excel_file = create_excel_file()
            
            if excel_file and os.path.exists(excel_file):
                # Перенаправляем на скачивание
                self.send_response(302)
                self.send_header('Location', '/download/gps_speed_data.xlsx')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                print(f'📊 Excel файл создан и готов к скачиванию: {excel_file}')
            else:
                # Показываем ошибку
                self.send_response(200)
                self.send_header('Content-type', 'text/html; charset=utf-8')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                
                html = f'''<!DOCTYPE html>
<html>
<head>
    <title>❌ Ошибка создания Excel - 69F СКОРОСТЬ</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 0; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; }}
        .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 10px 30px rgba(0,0,0,0.2); overflow: hidden; }}
        .header {{ background: linear-gradient(135deg, #dc3545 0%, #c82333 100%); color: white; padding: 30px; text-align: center; }}
        .content {{ padding: 30px; }}
        .error {{ background: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; padding: 15px; border-radius: 8px; margin-bottom: 20px; }}
        .back-link {{ display: inline-block; background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin-top: 20px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>❌ Ошибка создания Excel</h1>
            <p>Не удалось создать Excel файл</p>
        </div>
        <div class="content">
            <div class="error">
                <strong>Ошибка:</strong> Не удалось создать Excel файл с данными о скорости.
                <br><br>
                Возможные причины:
                <ul>
                    <li>Нет данных от устройств</li>
                    <li>Ошибка обработки файлов</li>
                    <li>Проблемы с библиотекой openpyxl</li>
                </ul>
            </div>
            <a href="/" class="back-link">← Вернуться к мониторингу</a>
        </div>
    </div>
</body>
</html>'''
                
                self.wfile.write(html.encode('utf-8'))
                print(f'❌ Ошибка создания Excel файла')
                
        except Exception as e:
            print(f'❌ Ошибка в handle_create_excel: {e}')
            self.send_error(500, "Internal server error")

    def do_GET(self):
        if self.path.startswith('/download/'):
            self.handle_file_download()
            return
            
        if self.path == '/cleanup':
            self.handle_cleanup()
            return
            
        if self.path == '/restart_tracking':
            self.handle_restart_tracking()
            return
            
        if self.path == '/create_excel':
            self.handle_create_excel()
            return
            
        self.send_response(200)
        self.send_header('Content-type', 'text/html; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        self.end_headers()
        
        # Читаем данные из файлов устройств
        device_files = []
        if os.path.exists(DATA_DIR):
            device_files = [f for f in os.listdir(DATA_DIR) if f.startswith('device_') and f.endswith('.txt') and not f.endswith('_log.txt')]
        
        devices_html = ""
        if not device_files:
            devices_html = '<div>Нет данных от устройств. Подключите Android приложения.</div>'
        else:
            for filename in sorted(device_files):
                filepath = os.path.join(DATA_DIR, filename)
                try:
                    with open(filepath, 'r') as f:
                        file_content = f.read().strip()
                    
                    # Парсим содержимое файла (скорость и время)
                    lines = file_content.split('\n')
                    if len(lines) >= 2:
                        speed = lines[0]
                        data_timestamp = lines[1]
                    else:
                        speed = file_content
                        data_timestamp = "Неизвестно"
                    
                    device_name = filename.replace('device_', '').replace('.txt', '').replace('_', ' ')
                    safe_name = device_name.replace(' ', '_')
                    
                    current_time = get_moscow_time()
                    last_update_time = datetime.fromtimestamp(os.path.getmtime(filepath), tz=pytz.timezone('Europe/Moscow'))
                    time_diff = (current_time - last_update_time).total_seconds()
                    
                    if time_diff > 10:
                        status_text = "🔴 Device not Tracking"
                        status_color = "#dc3545"
                        speed_display = "—"
                    else:
                        status_text = "🟢 Device Tracking"
                        status_color = "#28a745"
                        speed_display = f"{speed} км/ч"
                    
                    devices_html += f'''
                    <div style="background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 20px; margin-bottom: 15px;">
                        <h2>📱 Устройство: {device_name}</h2>
                        <div style="color: {status_color}; font-weight: bold; margin-bottom: 10px;">{status_text}</div>
                        <div style="font-size: 2em; font-weight: bold; color: #28a745; margin: 10px 0;">{speed_display}</div>
                        <div style="font-size: 0.9em; color: #6c757d; margin: 5px 0;">
                            ⏰ Последние данные: {data_timestamp} (МСК)
                        </div>
                        <div style="margin-top: 15px; display: flex; gap: 10px; flex-wrap: wrap;">
                            <a href="/download/device_{safe_name}.txt" style="color: #007bff; text-decoration: none; padding: 8px 12px; background: #e3f2fd; border-radius: 5px; font-size: 0.9em;">📄 Текущая скорость</a>
                            <a href="/download/device_{safe_name}_log.txt" style="color: #28a745; text-decoration: none; padding: 8px 12px; background: #e8f5e8; border-radius: 5px; font-size: 0.9em;">📊 История</a>
                        </div>
                        <div style="margin-top: 15px; padding: 10px; background: #fff; border: 1px solid #dee2e6; border-radius: 5px;">
                            <div style="font-size: 0.9em; color: #495057; margin-bottom: 8px;">📋 Копировать ссылки:</div>
                            <div style="display: flex; gap: 8px; flex-wrap: wrap;">
                                <button onclick="copyToClipboard('https://gps-speed-tracker.vercel.app/download/device_{safe_name}.txt')" 
                                        style="background: #007bff; color: white; border: none; padding: 6px 12px; border-radius: 4px; cursor: pointer; font-size: 0.8em;">
                                    📄 Скопировать ссылку на скорость
                                </button>
                                <button onclick="copyToClipboard('https://gps-speed-tracker.vercel.app/download/device_{safe_name}_log.txt')" 
                                        style="background: #28a745; color: white; border: none; padding: 6px 12px; border-radius: 4px; cursor: pointer; font-size: 0.8em;">
                                    📊 Скопировать ссылку на историю
                                </button>
                            </div>
                        </div>
                    </div>
                    '''
                except Exception as e:
                    devices_html += f'<div>❌ Ошибка чтения {filename}: {e}</div>'

        html_content = f'''<!DOCTYPE html>
<html>
<head>
    <title>⛵ 69F СКОРОСТЬ</title>
    <meta http-equiv="refresh" content="1">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 0; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 10px 30px rgba(0,0,0,0.2); overflow: hidden; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; }}
        .content {{ padding: 30px; }}
        .status {{ text-align: center; color: #6c757d; margin-bottom: 20px; }}
    </style>
    <script>
        function copyToClipboard(text) {{
            navigator.clipboard.writeText(text).then(function() {{
                // Показываем уведомление об успешном копировании
                const notification = document.createElement('div');
                notification.style.cssText = `
                    position: fixed;
                    top: 20px;
                    right: 20px;
                    background: #28a745;
                    color: white;
                    padding: 12px 20px;
                    border-radius: 5px;
                    font-size: 14px;
                    z-index: 1000;
                    box-shadow: 0 4px 8px rgba(0,0,0,0.2);
                `;
                notification.textContent = '✅ Ссылка скопирована в буфер обмена!';
                document.body.appendChild(notification);
                
                // Удаляем уведомление через 3 секунды
                setTimeout(() => {{
                    document.body.removeChild(notification);
                }}, 3000);
            }}).catch(function(err) {{
                console.error('Ошибка копирования: ', err);
                // Fallback для старых браузеров
                const textArea = document.createElement('textarea');
                textArea.value = text;
                document.body.appendChild(textArea);
                textArea.select();
                try {{
                    document.execCommand('copy');
                    alert('Ссылка скопирована в буфер обмена!');
                }} catch (err) {{
                    alert('Не удалось скопировать ссылку. Скопируйте вручную: ' + text);
                }}
                document.body.removeChild(textArea);
            }});
        }}
    </script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>⛵ 69F СКОРОСТЬ</h1>
            <p>Отслеживание скорости всех устройств</p>
            <div style="margin-top: 15px;">
                <a href="/cleanup" style="background: rgba(255,255,255,0.2); color: white; padding: 8px 16px; text-decoration: none; border-radius: 5px; font-size: 0.9em; margin-right: 10px;">🧹 Очистить старые данные</a>
                <a href="/restart_tracking" style="background: rgba(255,255,255,0.2); color: white; padding: 8px 16px; text-decoration: none; border-radius: 5px; font-size: 0.9em; margin-right: 10px;">🔄 Перезапустить Tracking</a>
                <a href="/download/all_devices.txt" style="background: rgba(255,255,255,0.2); color: white; padding: 8px 16px; text-decoration: none; border-radius: 5px; font-size: 0.9em; margin-right: 10px;">📥 Скачать все данные</a>
                <a href="/download/GPS-Speed-69F-v3.0-With-Remote-Restart.apk" style="background: rgba(255,255,255,0.2); color: white; padding: 8px 16px; text-decoration: none; border-radius: 5px; font-size: 0.9em;">📱 Скачать APK</a>
            </div>
        </div>
        <div class="content">
            <div class="status">Обновлено: {get_moscow_time().strftime('%Y-%m-%d %H:%M:%S')} (МСК)</div>
            {devices_html}
            
            <div style="background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 20px; margin-top: 30px;">
                <h2>📁 Прямые ссылки на файлы</h2>
                <div style="background: white; padding: 20px; border-radius: 8px;">
                    
                    <!-- APK файл -->
                    <div style="margin-bottom: 25px; padding: 15px; background: #e3f2fd; border-radius: 8px; border-left: 4px solid #2196f3;">
                        <h3 style="margin: 0 0 10px 0; color: #1976d2;">📱 Android приложение</h3>
                        <p style="margin: 5px 0; color: #666;">GPS Speed 69F v3.0 с удаленным перезапуском</p>
                        <a href="/download/GPS-Speed-69F-v3.0-With-Remote-Restart.apk" 
                           style="display: inline-block; background: #2196f3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold; margin-right: 10px;">
                            📥 Скачать APK
                        </a>
                        <span style="color: #666; font-size: 0.9em;">Размер: 4.7 MB</span>
                    </div>
                    
                    <!-- Общий лог -->
                    <div style="margin-bottom: 25px; padding: 15px; background: #f3e5f5; border-radius: 8px; border-left: 4px solid #9c27b0;">
                        <h3 style="margin: 0 0 10px 0; color: #7b1fa2;">📋 Общий лог всех устройств</h3>
                        <p style="margin: 5px 0; color: #666;">Все данные о скорости в одном файле</p>
                        <a href="/download/all_devices.txt" 
                           style="display: inline-block; background: #9c27b0; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold; margin-right: 10px;">
                            📥 Скачать лог
                        </a>
                        <span style="color: #666; font-size: 0.9em;">Формат: текст</span>
                    </div>
                    
                    <!-- Файлы устройств -->
                    <div style="margin-bottom: 20px;">
                        <h3 style="margin: 0 0 15px 0; color: #2e7d32;">📊 Файлы отдельных устройств</h3>
                        {self.get_device_links_html()}
                    </div>
                    
                    <!-- Excel файл -->
                    <div style="margin-bottom: 25px; padding: 15px; background: #e8f5e8; border-radius: 8px; border-left: 4px solid #4caf50;">
                        <h3 style="margin: 0 0 10px 0; color: #2e7d32;">📊 Excel отчет</h3>
                        <p style="margin: 5px 0; color: #666;">Excel файл обновляется в реальном времени при получении данных</p>
                        <a href="/download/gps_speed_data.xlsx" 
                           style="display: inline-block; background: #4caf50; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold; margin-right: 10px;">
                            📊 Скачать Excel файл
                        </a>
                        <span style="color: #666; font-size: 0.9em;">Устройство, скорость, время (чистые данные без цветов)</span>
                    </div>
                    
                    <!-- Служебные файлы -->
                    <div style="margin-top: 25px; padding: 15px; background: #fff3e0; border-radius: 8px; border-left: 4px solid #ff9800;">
                        <h3 style="margin: 0 0 10px 0; color: #f57c00;">🔧 Служебные файлы</h3>
                        <div style="margin: 10px 0;">
                            <a href="/download/restart_signal.txt" 
                               style="display: inline-block; background: #ff9800; color: white; padding: 8px 16px; text-decoration: none; border-radius: 5px; font-weight: bold; margin-right: 10px; margin-bottom: 5px;">
                                🔄 Файл-сигнал перезапуска
                            </a>
                            <span style="color: #666; font-size: 0.9em;">Для удаленного перезапуска tracking</span>
                        </div>
                    </div>
                    
                </div>
            </div>
        </div>
    </div>
</body>
</html>'''
        
        self.wfile.write(html_content.encode('utf-8'))

    def get_device_links_html(self):
        """Генерирует HTML со ссылками на файлы устройств"""
        try:
            device_files = []
            if os.path.exists(DATA_DIR):
                device_files = [f for f in os.listdir(DATA_DIR) if f.startswith('device_') and f.endswith('.txt') and not f.endswith('_log.txt')]

            if not device_files:
                return '<div style="color: #6c757d; font-style: italic; padding: 20px; text-align: center; background: #f8f9fa; border-radius: 5px;">Нет файлов устройств</div>'

            links_html = ""
            for filename in sorted(device_files):
                device_name = filename.replace('device_', '').replace('.txt', '').replace('_', ' ')
                safe_name = device_name.replace(' ', '_')

                # Получаем информацию о файле
                device_file = os.path.join(DATA_DIR, filename)
                log_file = os.path.join(DATA_DIR, f'device_{safe_name}_log.txt')
                
                device_size = os.path.getsize(device_file) if os.path.exists(device_file) else 0
                log_size = os.path.getsize(log_file) if os.path.exists(log_file) else 0

                links_html += f'''
                <div style="margin-bottom: 20px; padding: 15px; background: #f8f9fa; border-radius: 8px; border-left: 4px solid #4caf50;">
                    <h4 style="margin: 0 0 10px 0; color: #2e7d32;">🚤 {device_name}</h4>
                    <div style="display: flex; gap: 10px; flex-wrap: wrap;">
                        <a href="/download/device_{safe_name}.txt" 
                           style="display: inline-block; background: #4caf50; color: white; padding: 8px 16px; text-decoration: none; border-radius: 5px; font-weight: bold; font-size: 0.9em;">
                            📄 Текущая скорость
                        </a>
                        <a href="/download/device_{safe_name}_log.txt" 
                           style="display: inline-block; background: #8bc34a; color: white; padding: 8px 16px; text-decoration: none; border-radius: 5px; font-weight: bold; font-size: 0.9em;">
                            📊 История данных
                        </a>
                    </div>
                    <div style="margin-top: 8px; color: #666; font-size: 0.8em;">
                        Размеры: скорость {device_size} байт, история {log_size} байт
                    </div>
                </div>
                '''

            return links_html
        except Exception as e:
            return f'<div style="color: #dc3545; padding: 15px; background: #f8d7da; border-radius: 5px;">Ошибка: {e}</div>'

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, X-Device-Name')
        self.end_headers()
